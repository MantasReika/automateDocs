# ~*~ coding: utf-8 ~*~

import logging
from openpyxl import load_workbook  # Workbook
import datetime
import os

class Summary:
    wb = ""
    ws = ""
   

    def __init__(self, fileName, currentMonth):
        self.currentMonth = currentMonth
        self.fileName = fileName
        
        self.startingRow = 6
        self.lastRow = 2500 + 1
        
        self.currentRowId = 0
        self.personFirstRowId = 0
        self.relativeCol = 11 + ((currentMonth - 1 ) * 7)

        self.fileName = fileName.split(".xlsx")[0]
        if not os.path.isfile(self.fileName + ".xlsx"):
            raise FileNotFoundError("summary file='{}' does not exist at path='{}'".format(self.fileName + ".xlsx", os.getcwd()))
      
        self.wb = load_workbook(self.fileName + ".xlsx")
        self.ws = self.wb.active

    def isValidStructuredDocument(self):
        # Check header row, months collumns contain 7 subcollumns

        return True

    def isValidRow(self):
        # Check current row data structures, is each month valid
        # if self.ws.cell(row=self.currentRowId, column=self.currentMonth + self.fromDateId).value is None:
        #     return False
        # if self.ws.cell(row=self.currentRowId, column=self.currentMonth + self.toDateId).value is None:
        #     return False
        
        return True

    def isCellDate(self, cell):
        if cell.value == None or type(cell.value) is not datetime.datetime:
            return False
        return True


    def isCellLastDayOfMonth(self, cell):
        if not self.isCellDate(cell):
            return False

        # cellDt = datetime.datetime.strptime(cell.value, '%Y-%m-%d')
        cellDt = cell.value
        
        nextMonth = cellDt.replace(day=28) + datetime.timedelta(days=4)
        lastDayOfMonth = nextMonth - datetime.timedelta(days=nextMonth.day)
        if lastDayOfMonth.day == cellDt.day:
            return True
        else:
            return False

    def lastDayOfMonth(self, dtObj):
        if dtObj == None or type(dtObj) is not datetime.datetime:
            return None
        
        next_month = dtObj.replace(day=28) + datetime.timedelta(days=4)
        return next_month - datetime.timedelta(days=next_month.day)
    
    def nextMonthLastDay(self, cell):
        if not self.isCellDate(cell):
            return None
        cellDt = cell.value
        next_month = cellDt.replace(day=28) + datetime.timedelta(days=4)
        return self.lastDayOfMonth(next_month)

    def nextMonthFirstDay(self, cell):
        if not self.isCellDate(cell):
            return None
        cellDt = cell.value
        next_month = cellDt.replace(day=28) + datetime.timedelta(days=4)
        return next_month.replace(day=1)

    def initRowCells(self):
        
        self.fromCell_1 = self.ws.cell(row=self.currentRowId, column=self.relativeCol+0)
        self.fromCell_2 = self.ws.cell(row=self.currentRowId, column=self.relativeCol+1)
        self.fromCell_5 = self.ws.cell(row=self.currentRowId, column=self.relativeCol+4)
        self.fromCell_6 = self.ws.cell(row=self.currentRowId, column=self.relativeCol+5)
        
        self.toCell_1 = self.ws.cell(row=self.personFirstRowId, column=self.relativeCol+7)
        self.toCell_2 = self.ws.cell(row=self.personFirstRowId, column=self.relativeCol+8)
        self.toCell_5 = self.ws.cell(row=self.personFirstRowId, column=self.relativeCol+11)
        self.toCell_6 = self.ws.cell(row=self.personFirstRowId, column=self.relativeCol+12)

    def resetCells(self):
        self.lomCell_1 = None
        self.lomCell_2 = None
        self.lomCell_5 = None
        self.lomCell_6 = None

    def checkRowCells(self):
#        print("ROW id: {}, self.checkRowCells is last day of month: {}, date: {}".format(self.currentRowId, self.isCellLastDayOfMonth(self.fromCell_6), self.fromCell_6.value))
        if self.isCellLastDayOfMonth(self.fromCell_6):
            self.lomCell_1 = self.fromCell_1
            self.lomCell_2 = self.fromCell_2
            self.lomCell_5 = self.nextMonthFirstDay(self.fromCell_5)
            self.lomCell_6 = self.nextMonthLastDay(self.fromCell_6)

    def copyDataToNextMonth(self):
        if self.lomCell_1 is not None:
#            print("ROW id: {}, self.copyDataToNextMonth".format(self.currentRowId))
            # copy 1 st , 2 nd and 5 th and 6 th columns with new month dates
            self.toCell_1.value = self.lomCell_1.value
            self.toCell_2.value = self.lomCell_2.value

            self.toCell_5.value = self.lomCell_5
            self.toCell_6.value = self.lomCell_6

    def processRows(self):
        # for row in ws.iter_rows(min_row=startingRow, max_row=lastRow, min_col=minCol, max_col=maxCol):
        rowCnt = 0
        self.resetCells()
        for row in range(self.startingRow, self.lastRow):
            rowCnt += 1
            self.currentRowId = row
            if rowCnt == 1:
                self.personFirstRowId = row

            self.initRowCells()
            self.checkRowCells()

            if rowCnt > 3:
                rowCnt = 0
                self.processMonth(self.currentMonth)
                self.resetCells()
                

    def processMonth(self, month):        
        self.copyDataToNextMonth()

    def saveFile(self):
        self.newFileName = self.fileName.strip(".xlsx") + "_generated_month-{}.xlsx".format(self.currentMonth)
        self.wb.save(self.newFileName)
        return
"""
if __name__ == "__main__":
    fileName = "Naujausia-kom.-suvestinė-2019m.xlsx"
    currentMonth = int(input("Enter month from which to copy: "))
    summ = Summary(fileName, currentMonth)
    summ.isValidStructuredDocument()
    summ.processRows()
    summ.saveFile()

    print("Done.")
"""