import logging
import copy
import datetime
import json
import os

import docx
import openpyxl
from docx import Document
from openpyxl import load_workbook

from autoSecondment.generatorUtils.Constants import Constants as SecondmentConstants
from autoSecondment.generatorUtils.secondmentUtils import (createSecondmentDocument, indexParagrahps, indexRuns, mapCodeToCountry, calculateLastWorkingDate)
from autoSecondment.generatorUtils.PersonRecord import PersonRecord

class SummaryReader:
    """
    Should open xlsx summary file
    and read data from selected month
    and return list of Person objects
    """

    wb = None
    ws = None
    currentMonth = None
    fileName = None

    def __init__(self):
        
        self.startingRow = 6
        self.lastRow = 1850 + 1
        
        self.currentRowId = 0
        self.personFirstRowId = 0
              
        self.personRecords = []
    
    def getPersonRecords(self):
        return self.personRecords

    def setSummaryRelativeCollumn(self, relativeToMonth):
        if type(relativeToMonth) != int and (relativeToMonth < 1 or relativeToMonth > 12):
            raise Exception("summary month='{}' value passed is invalid".format(month))

        self.relativeCol = 10 + ((relativeToMonth - 1 ) * 7)
    
    def getSummaryFileName(self):
        return self.fileName + ".xlsx"

    def openSummaryFile(self, fileName, month):
        if type(month) != int and (month < 1 or month > 12):
            raise Exception("summary month='{}' value passed is invalid".format(month))

        self.currentMonth = month
        
        self.fileName = fileName.split('.xlsx')[0]
        if not os.path.isfile(self.getSummaryFileName()):
            raise FileNotFoundError("summary file='{}' does not exist at path='{}'".format(self.getSummaryFileName(), os.getcwd()))

        self.wb = load_workbook(filename=self.getSummaryFileName())
        self.ws = self.wb.active

    def getCellValue(self, rowId, collumnId):
        try:
            return self.ws.cell(row=rowId, column=collumnId).value
        except Exception as e:
            logging.warning("getCellValue Exception: {}".format(e))
            return None
    
    def getCellDate(self, rowId, collumnId):
        cell = self.getCellValue(rowId, collumnId)
        if not self.isCellDate(cell):
            return None
            
        return cell.strftime("%Y-%m-%d")
        
    def isCellDate(self, cell):
        if cell == None or type(cell) is not datetime.datetime:
            return False
        return True

    def isEmptyRow(self, rowId):
        if rowId > self.lastRow:
            logging.warning("Last summary row loaded, stopping loading; rowId > self.lastRow; {} > {}".format(rowId, self.lastRow))
            return True
        try:
            if self.getCellValue(rowId, 2) == None:
                return True
        except Exception as e:
            logging.warning("While checking summary row, rowId='{}', stopping loading; error: {}".format(rowId, e))
            return True
        return False
    
    def readInPersonRecords(self):
        """
        For current month read in all person records
        """
        self.currentRowId = self.startingRow
        previousFullName = ""

        while not self.isEmptyRow(self.currentRowId):
            fullName = self.getCellValue(self.currentRowId, 2)
            
            # take only if new person 
            if fullName != previousFullName:
                workDistrict = str(self.getCellValue(self.currentRowId, 4)).upper()
                proffesion = str(self.getCellValue(self.currentRowId, 5)).lower()
            
            docNr = self.getCellValue(self.currentRowId, self.relativeCol + 4)
            sentCountry = mapCodeToCountry(self.getCellValue(self.currentRowId, self.relativeCol + 1))
            dateFrom = self.getCellDate(self.currentRowId, self.relativeCol + 5)
            dateTo = self.getCellDate(self.currentRowId, self.relativeCol + 6)
            docDate = calculateLastWorkingDate(dateFrom)
             
            self.savePersonData(fullName, workDistrict, proffesion, docNr, docDate, sentCountry, dateFrom, dateTo)
            
            self.currentRowId += 1
            previousFullName = fullName


    def savePersonData(self, fullName, workDistrict, proffesion, docNr, docDate, sentCountry, dateFrom, dateTo):
        record = PersonRecord(fullName, workDistrict, proffesion, docNr, docDate, sentCountry, dateFrom, dateTo)
        
        if record.hasAllFieldsFilled():
            self.personRecords.append(record)
            logging.debug([self.currentRowId, record.hasAllFieldsFilled(), record])
    
    def readCollumns(self):
        for i in range(1, 20):
            logging.debug("row: 1, col: {}, val: '{}'".format(i, self.getCellValue(6, i)))
